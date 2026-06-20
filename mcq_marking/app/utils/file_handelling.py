import json
import os
from PIL import Image
import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from app.storage.nfs_storage import NFSStorage
from io import BytesIO
import logging
import csv
from io import StringIO

logger = logging.getLogger(__name__)

def save_image(path, image):
    """
    Save image to NFS storage
    
    Args:
        path: Relative path within the file_type directory
        image: PIL Image or OpenCV image
    """
    # Validate image before processing
    if image is None:
        raise ValueError("Cannot save None image")
    
    # Convert image to bytes
    if hasattr(image, 'save'):            # PIL Image
        img_bytes = BytesIO()
        image.save(img_bytes, format='PNG')
        image_bytes = img_bytes.getvalue()
    else:
        # OpenCV image (numpy array)
        # Check if image is empty
        if not isinstance(image, np.ndarray) or image.size == 0:
            raise ValueError("Cannot save empty OpenCV image")
        
        success, img_bytes = cv2.imencode('.png', image)
        if success:
            image_bytes = img_bytes.tobytes()
        else:
            raise ValueError("Failed to encode OpenCV image")
        
    # Save to NFS storage
    nfs = NFSStorage()
    nfs.save_file(image_bytes, path)
    

def save_image_using_folder_and_filename(folder_path, filename, image):
    """Save image using folder and filename with NFS support"""
    save_image(os.path.join(folder_path, filename), image)

def read_image(path, convert_to_grayscale=False, test=False):
    """
    Read image from NFS storage or local file system
    
    Args:
        path: Relative path within the file_type directory (for NFS) or absolute/relative path (for local)
        convert_to_grayscale: Whether to convert to grayscale
        test: If True, read from local file system instead of NFS
        
    Returns:
        PIL Image object
    """
    if test:
        # Read from local file system
        image = Image.open(path)
    else:
        # Read from NFS storage
        nfs = NFSStorage()
        image_bytes = nfs.get_file(path)
        image = Image.open(BytesIO(image_bytes))
    
    if convert_to_grayscale:
        return image.convert('L')
    return image

class NumpyEncoder(json.JSONEncoder):
    """Custom JSON encoder for numpy data types"""
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        return super(NumpyEncoder, self).default(obj)

def save_json(data, path):
    """
    Save JSON data to NFS storage
    
    Args:
        data: Data to save as JSON
        path: Relative path within the file_type directory
    """
    nfs = NFSStorage()
    json_bytes = json.dumps(data, indent=2, cls=NumpyEncoder).encode('utf-8')
    nfs.save_file(json_bytes, path)

def read_json(path):
    """
    Read JSON data from NFS storage
    
    Args:
        path: Relative path within the file_type directory
        
    Returns:
        Parsed JSON data
    """
    nfs = NFSStorage()
    json_bytes = nfs.get_file(path)
    return json.loads(json_bytes.decode('utf-8'))

def read_answer_sheet_paths(folder_path):
    """
    List all answer sheet file paths in a given NFS folder.

    Args:
        folder_path: Relative path within the file_type directory (e.g., 'uploads/answer_sheets')

    Returns:
        Sorted list of file paths (relative to file_type/folder_path)
    """
    nfs = NFSStorage()
    # List files in the given subdirectory using NFS
    file_list = nfs.list_files(directory=folder_path)
    # Return sorted list of relative paths (folder_path/filename)
    return sorted(file_list)

def file_exists(path):
    """
    Check if file exists in NFS storage
    
    Args:
        path: Relative path within the file_type directory
    """
    if path is None:
        return False
    nfs = NFSStorage()
    return nfs.file_exists(path)

def get_spreadsheet(path, title: str):
    """
    Get or create a spreadsheet from NFS storage
    
    Args:
        path: Relative path within the file_type directory
        title: Name of the sheet to get or create
        
    Returns:
        Tuple of (workbook, sheet)
    """
    nfs = NFSStorage()
    workbook = None
    sheet = None
    
    try:
        # Try to get existing file from NFS
        file_bytes = nfs.get_file(path)
        workbook = load_workbook(BytesIO(file_bytes))
        
        # Check if a sheet with the given title exists
        if title in workbook.sheetnames:
            sheet = workbook[title]
        else:
            sheet = workbook.create_sheet(title)
    except FileNotFoundError:
        # File doesn't exist, create new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title
    
    return workbook, sheet

def save_spreadsheet(path, workbook):
    """
    Save spreadsheet to NFS storage
    
    Args:
        path: Relative path within the file_type directory
        workbook: OpenPyXL workbook object to save
    """
    nfs = NFSStorage()
    
    # Save workbook to BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    # Save to NFS storage
    nfs.save_file(buffer.getvalue(), path)

# Functions for getting the index numbers list from a file path(supports csv or xlsx)
def get_column_from_file(file_path, column_name):
    """
    Get a specific column from a CSV or XLSX file stored in NFS storage.
    
    Args:
        file_path: Relative path within the file_type directory
        column_name: Name of the column to extract
        
    Returns:
        List of values in the specified column
    """
    nfs = NFSStorage()
    _, ext = os.path.splitext(file_path)
    
    if ext.lower() == '.csv':
        # Read CSV file
        file_bytes = nfs.get_file(file_path)
        file_str = file_bytes.decode('utf-8')
        csv_reader = csv.DictReader(StringIO(file_str))
        if column_name not in csv_reader.fieldnames:
            raise ValueError(f"Column '{column_name}' not found in CSV file")
        return [row[column_name] for row in csv_reader if row[column_name] is not None]

    
    elif ext.lower() in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
        # Read XLSX file
        file_bytes = nfs.get_file(file_path)
        workbook = load_workbook(BytesIO(file_bytes), read_only=True)
        sheet = workbook.active
        
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        
        if column_name not in header:
            raise ValueError(f"Column '{column_name}' not found in XLSX file")
        
        col_index = header.index(column_name) + 1  # openpyxl is 1-indexed
        return [row[col_index - 1].value for row in sheet.iter_rows(min_row=2) if row[col_index - 1].value is not None]
    
    else:
        raise ValueError("Unsupported file format. Only CSV and XLSX are supported.")