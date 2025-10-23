import uuid

import pandas as pd
from fastapi import APIRouter, HTTPException, WebSocket, WebSocketDisconnect, Request, Response
from starlette.websockets import WebSocketState
from sqlalchemy import select
import logging
from io import BytesIO
import zipfile
import os
from pathlib import Path

from sqlalchemy.orm import selectinload
from openpyxl import load_workbook

from app.models.marking_job import MarkingJob, MarkingJobStatus
from app.models.template import Template
from app.schemas.marking import MarkingCreateMetadata, MarkingResponse, MarkingAttachAnswerSheets, MarkingResponseBasic, ProgressRequest, ProgressResponse, ResultsData, UpdateResultRequest, MarkingAttachIndexList
from app.schemas.marking import UpdateMarkingSchemeConfigRequest
from app.database import get_async_db
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import Depends
from typing import List
from app.api.deps import get_websocket_manager
from app.websocket import WebSocketManager
from app.middleware.websocket_auth import authorize_websocket
from app.models.user import UserRoles

from app.queue import submit_marking_job, submit_marking_scheme_config_job
from app.storage.shared_storage import SharedStorage
from app.middleware.authorization import require_non_super_admin
from app.middleware.websocket_auth import authorize_websocket
from app.models.user import UserRoles
import json

router = APIRouter(prefix="/api/markings", tags=['markings'])

logger = logging.getLogger(__name__)


def _get_marking_response(marking: MarkingJob) -> MarkingResponse:
    return MarkingResponse(
            id=marking.id,
            name=marking.name,
            description=marking.description,
            status=marking.status,
            priority=marking.priority,
            template_id=marking.template_id,
            marking_scheme_id=marking.marking_scheme_id,
            marking_config_id=marking.marking_config_id,
            answer_sheets_folder_id=marking.answer_sheets_folder_id,
            result_sheet_file_id=marking.result_sheet_file_id,
            index_list_file_id=marking.index_list_file_id,
            save_intermediate_results=marking.save_intermediate_results,
            total_answer_sheets=marking.total_answer_sheets,
            processed_answer_sheets=marking.processed_answer_sheets,
            failed_answer_sheets=marking.failed_answer_sheets,
            processing_started_at=marking.processing_started_at,
            processing_completed_at=marking.processing_completed_at,
            error_message=marking.error_message,
            error_details=marking.error_details,
            results_summary=marking.results_summary,
            created_at=marking.created_at,
            updated_at=marking.updated_at,
            created_by=marking.created_by
        )
        



@router.get('/', response_model=List[MarkingResponseBasic])
@require_non_super_admin(require_admin_verified=True)
async def list_markings(
    request: Request,
    db: AsyncSession = Depends(get_async_db)
):
    """List all markings"""
    # Get user from token for ownership filtering
    user_id = request.state.current_user.id
    try:
      markings = await db.execute(select(MarkingJob).options(selectinload(MarkingJob.template)).where(MarkingJob.created_by == user_id).order_by(MarkingJob.created_at.desc()))
      markings = markings.scalars().all()
      return [
          MarkingResponseBasic(
              id=marking.id,
              name=marking.name,
              status=marking.status,
              priority=marking.priority,
              template_name=marking.template.name,
              save_intermediate_results=marking.save_intermediate_results,
              total_answer_sheets=marking.total_answer_sheets,
              processed_answer_sheets=marking.processed_answer_sheets,
              created_at=marking.created_at,
              updated_at=marking.updated_at,
              created_by=marking.created_by
          )
          for marking in markings
      ]
    except Exception as e:
        logger.error(f"Failed to list markings: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to list markings")


@router.websocket("/progress")
async def progress_websocket(
    websocket: WebSocket,
    websocket_manager: WebSocketManager = Depends(get_websocket_manager),
    db: AsyncSession = Depends(get_async_db)
):
    """Get the progress for a marking job"""
    marking_job_ids = []
    
    try:
        # Accept WebSocket connection first (required before authorization can close it)
        await websocket.accept()
        
        user = await authorize_websocket(
            websocket,
            allowed_roles=[UserRoles.BASIC, UserRoles.FACULTYADMIN],
            require_admin_verified=True
        )
        user_id = user.id

        # Receive the marking job IDs from the client
        data = await websocket.receive_json()
        
        # Validate the received data
        if "marking_job_ids" not in data:
            await websocket.send_json({
                "status": "error",
                "message": "Missing marking_job_ids in request data"
            })
            await websocket.close()
            return
        
        marking_job_ids = data["marking_job_ids"]
        
        # Validate that all marking jobs belong to the user
        for marking_job_id in marking_job_ids:
            result = await db.execute(
                select(MarkingJob).where(
                    MarkingJob.id == marking_job_id,
                    MarkingJob.created_by == user_id
                )
            )
            marking_job = result.scalar_one_or_none()
            
            if not marking_job:
                await websocket.send_json({
                    "status": "error",
                    "message": f"Marking job {marking_job_id} not found or access denied"
                })
                await websocket.close()
                return
        
        # Register to all marking jobs (websocket already accepted)
        for marking_job_id in marking_job_ids:
            await websocket_manager.register_marking_job(str(marking_job_id), websocket)
        
        # Send connection confirmation
        await websocket.send_json({
            "status": "connected",
            "message": "Connected to progress websocket"
        })
        
        # Keep the connection open to receive progress updates
        try:
            while True:
                await websocket.receive_text()
        except WebSocketDisconnect:
            logger.info(f"WebSocket disconnected for marking jobs {marking_job_ids}")
            
    except WebSocketDisconnect:
        logger.info(f"WebSocket disconnected")
    except Exception as e:
        logger.error(f"Failed to connect to progress websocket: {str(e)}")
        if websocket.client_state == WebSocketState.CONNECTED:
            try:
                await websocket.send_json({
                    "status": "error",
                    "message": f"Failed to connect to progress websocket: {str(e)}"
                })
            except Exception:
                pass
    finally:
        # Disconnect from all marking jobs
        for marking_job_id in marking_job_ids:
            try:
                await websocket_manager.disconnect_marking_job(str(marking_job_id), websocket)
            except Exception as e:
                logger.warning(f"Failed to disconnect from marking job {marking_job_id}: {e}")

@router.get("/{marking_job_id}", response_model=MarkingResponse)
@require_non_super_admin(require_admin_verified=True)
async def get_marking(
    request: Request,
    marking_job_id: int,
    db: AsyncSession = Depends(get_async_db)
):
    """Get a single marking job by ID"""
    # Get user from token for ownership validation
    user_id = request.state.current_user.id
    try:
        result = await db.execute(
            select(MarkingJob).where(
                MarkingJob.id == marking_job_id,
                MarkingJob.created_by == user_id
            )
        )
        marking = result.scalar_one_or_none()
        
        if not marking:
            raise HTTPException(status_code=404, detail="Marking job not found")
        
        return _get_marking_response(marking)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get marking job {marking_job_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to get marking job")

@router.post("/", response_model=MarkingResponse, status_code=201)
@require_non_super_admin(require_admin_verified=True)
async def create_marking(
    request: Request,
    marking: MarkingCreateMetadata,
    db: AsyncSession = Depends(get_async_db)
):
    """Create a new marking"""
    # Get user from token for ownership tracking
    user_id = request.state.current_user.id
    try:
        marking_record = MarkingJob(
          name=marking.name,
          description=marking.description,
          status=MarkingJobStatus.INITIALIZED,
          template_id=marking.template_id,
          save_intermediate_results=marking.save_intermediate_results,
          priority=marking.priority,
          created_by=user_id
        )
        db.add(marking_record)
        await db.commit()
        await db.refresh(marking_record)

        # Generate paths for future use but don't enqueue yet
        random_id = str(uuid.uuid4())[:8]
        marking_record.intermediate_results_path = f"intermediate/markings/{user_id}/{marking_record.id}_{random_id}_intermediate/"
        marking_record.result_sheet_file_path = f"results/markings/{user_id}/{marking_record.id}_{random_id}_output.xlsx"

        await db.commit()
        await db.refresh(marking_record)
        marking_response = _get_marking_response(marking_record)
        return marking_response
    except Exception as e:
        logger.error(f"Failed to create marking: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create marking")


@router.put("/{marking_job_id}", response_model=MarkingResponse)
@require_non_super_admin(require_admin_verified=True)
async def update_marking(
    request: Request,
    marking_job_id: int,
    marking: MarkingCreateMetadata,
    db: AsyncSession = Depends(get_async_db)
):
    """Update a marking"""
    # Get user from token for ownership validation
    user_id = request.state.current_user.id
    try:
        marking_record = await db.get(MarkingJob, marking_job_id)
        if not marking_record:
            raise HTTPException(status_code=404, detail="Marking job not found")
        
        marking_record.name = marking.name
        marking_record.description = marking.description
        marking_record.template_id = marking.template_id
        marking_record.save_intermediate_results = marking.save_intermediate_results
        marking_record.priority = marking.priority
        await db.commit()
        await db.refresh(marking_record)
        return _get_marking_response(marking_record)
    except Exception as e:
        logger.error(f"Failed to update marking: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update marking")


@router.websocket("/{marking_job_id}/configure-marking-scheme")
async def configure_marking_scheme_websocket(
    marking_job_id: int,
    websocket: WebSocket,
    db: AsyncSession = Depends(get_async_db),
    websocket_manager: WebSocketManager = Depends(get_websocket_manager)
):
    """Attach marking scheme and enqueue config job via WebSocket"""
    
    logger.info(f"WebSocket endpoint called for marking job {marking_job_id}")
    
    try:
        # Accept WebSocket connection first (required before authorization can close it)
        await websocket.accept()

        user = await authorize_websocket(
            websocket,
            allowed_roles=[UserRoles.BASIC, UserRoles.FACULTYADMIN],
            require_admin_verified=True
        )
        user_id = user.id
        
        logger.info(f"WebSocket connection established for marking job {marking_job_id}")
        
        # Register with WebSocket manager after accepting connection
        await websocket_manager.register_marking_scheme_config(str(marking_job_id), websocket)
        
        # Wait for the marking scheme data from the client
        
        data = await websocket.receive_json()
        logger.info(f"Received marking scheme data: {data}")
        
        # Validate the received data
        if "marking_scheme_id" not in data:
            await websocket.send_json({
                "status": "error",
                "message": "Missing marking_scheme_id in request data"
            })
            await websocket_manager.disconnect_marking_scheme_config(str(marking_job_id), websocket)
            return
        
        marking_scheme_id = data["marking_scheme_id"]
        
        # Get the marking job
        result = await db.execute(
            select(MarkingJob).where(
                MarkingJob.id == marking_job_id,
                MarkingJob.created_by == user_id
            )
        )
        marking = result.scalar_one_or_none()
        
        if not marking:
            await websocket.send_json({
                "status": "error",
                "message": "Marking job not found"
            })
            await websocket_manager.disconnect_marking_scheme_config(str(marking_job_id), websocket)
            return
        
        # Update marking scheme ID and generate marking config file path
        marking.marking_scheme_id = marking_scheme_id
        
        # Generate marking config file path
        random_id = str(uuid.uuid4())[:8]
        marking.marking_config_file_path = f"intermediate/markings/{user_id}/{marking.id}_{random_id}_marking_config.json"
        
        await db.commit()
        await db.refresh(marking)
        
        # Submit marking scheme configuration job to the queue
        try:
            logger.info(f"Submitting marking scheme config job for marking job {marking_job_id}")
            success = await submit_marking_scheme_config_job(marking_job_id, db)
            if success:
                marking.status = MarkingJobStatus.QUEUED
                await db.commit()
                await db.refresh(marking)
                logger.info(f"Successfully submitted marking scheme config job {marking_job_id} to queue")
                
                # Send success response
                await websocket.send_json({
                    "status": "queued",
                    "message": "Marking scheme configuration job queued successfully"
                })
                logger.info(f"WebSocket connection kept open for marking job {marking_job_id} to receive progress updates")
                
                # Keep connection alive - wait for disconnection
                try:
                    await websocket.receive_text()
                except Exception:
                    pass
            else:
                logger.error(f"Failed to submit marking scheme config job {marking_job_id} to queue")
                await websocket.send_json({
                    "status": "error",
                    "message": "Failed to submit marking scheme configuration job to queue"
                })
                await websocket_manager.disconnect_marking_scheme_config(str(marking_job_id), websocket)
                return
        except Exception as e:
            logger.error(f"Error submitting marking scheme config job {marking_job_id}: {e}")
            marking.status = MarkingJobStatus.FAILED
            marking.error_message = f"Failed to submit job to queue: {str(e)}"
            await db.commit()
            await db.refresh(marking)
            await websocket.send_json({
                "status": "error",
                "message": f"Failed to submit marking scheme configuration job: {str(e)}"
            })
            await websocket_manager.disconnect_marking_scheme_config(str(marking_job_id), websocket)
            return
            
    except Exception as e:
        logger.error(f"Failed to configure marking job {marking_job_id}: {str(e)}")
        logger.error(f"Exception type: {type(e).__name__}")
        logger.error(f"Exception details: {str(e)}")
        
        # Only try to send error message if WebSocket is still connected
        if websocket.client_state == WebSocketState.CONNECTED:
            try:
                await websocket.send_json({
                    "status": "error",
                    "message": f"Failed to configure marking job: {str(e)}"
                })
            except Exception as send_error:
                logger.warning(f"Failed to send error message via WebSocket: {send_error}")
        
        # Try to disconnect gracefully
        try:
            await websocket_manager.disconnect_marking_scheme_config(str(marking_job_id), websocket)
        except Exception as disconnect_error:
            logger.warning(f"Failed to disconnect WebSocket cleanly: {disconnect_error}")

@router.post("/{marking_job_id}/update-marking-scheme-config", response_model=MarkingResponse)
@require_non_super_admin(require_admin_verified=True)
async def update_marking_scheme_config(
    request_obj: Request,
    marking_job_id: int,
    request: UpdateMarkingSchemeConfigRequest,
    db: AsyncSession = Depends(get_async_db)
):
    """Update the marking scheme configuration for a job"""
    # Get user from token for ownership validation
    user_id = request_obj.state.current_user.id
    try:
        # Get the marking job
        result = await db.execute(
            select(MarkingJob).where(
                MarkingJob.id == marking_job_id,
                MarkingJob.created_by == user_id
            )
        )
        marking = result.scalar_one_or_none()
        
        if not marking:
            raise HTTPException(status_code=404, detail="Marking job not found")
        
        # Validate that the job is in the correct status to update marking scheme config
        if marking.status in [MarkingJobStatus.INITIALIZED]:
            raise HTTPException(
                status_code=400, 
                detail="Job must be in marking scheme configured or higher to verify"
            )
        
        # Validate that the job has a marking config file path
        if not marking.marking_config_file_path:
            raise HTTPException(
                status_code=400, 
                detail="Marking scheme must be configured first before updating"
            )
        
        if not request.isUpdated:
            marking.status = MarkingJobStatus.MARKING_SCHEME_VERIFIED
        else:

        
            # Update the marking config file with the new configuration
            try:
                # Save the updated configuration to the storage
                storage = SharedStorage()
                config_json = json.dumps(request.marking_scheme_config, indent=2)
                await storage.save_file(
                    config_json.encode('utf-8'),
                    marking.marking_config_file_path
                )
                
                marking.status = MarkingJobStatus.MARKING_SCHEME_VERIFIED
                
                logger.info(f"Successfully updated marking scheme config for job {marking_job_id}")
                
            except Exception as e:
                logger.error(f"Failed to save updated marking scheme config: {str(e)}")
                raise HTTPException(
                    status_code=500, 
                    detail=f"Failed to save updated configuration: {str(e)}"
                )
            
                
        await db.commit()
        await db.refresh(marking)
        
        return _get_marking_response(marking)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update marking scheme config for job {marking_job_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update marking scheme configuration")


@router.post("/{marking_job_id}/attach-answer-sheets", response_model=MarkingResponse)
@require_non_super_admin(require_admin_verified=True)
async def attach_answer_sheets(
    request: Request,
    marking_job_id: int,
    sheets_data: MarkingAttachAnswerSheets,
    db: AsyncSession = Depends(get_async_db)
):
    """Attach answer sheets folder to marking job"""
    # Get user from token for ownership validation
    user_id = request.state.current_user.id
    try:
        # Get the marking job
        result = await db.execute(
            select(MarkingJob).where(
                MarkingJob.id == marking_job_id,
                MarkingJob.created_by == user_id
            )
        )
        marking = result.scalar_one_or_none()
        
        if not marking:
            raise HTTPException(status_code=404, detail="Marking job not found")
        
        if marking.status not in [MarkingJobStatus.MARKING_SCHEME_VERIFIED, MarkingJobStatus.ANSWER_SHEETS_ATTACHED, MarkingJobStatus.COMPLETED]:
            raise HTTPException(status_code=400, detail="Job must be in marking scheme verified status or higher to attach answer sheets")
        
        # Update answer sheets folder ID
        marking.answer_sheets_folder_id = sheets_data.answer_sheets_folder_id
        marking.status = MarkingJobStatus.ANSWER_SHEETS_ATTACHED
        await db.commit()
        await db.refresh(marking)
        
        logger.info(f"Answer sheets attached to job {marking_job_id}")
        
        return _get_marking_response(marking)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to attach answer sheets to job {marking_job_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to attach answer sheets")

@router.post("/{marking_job_id}/attach-index-list", response_model=MarkingResponse)
async def attach_index_list(
    marking_job_id: int,
    index_list_data: MarkingAttachIndexList,
    db: AsyncSession = Depends(get_async_db)
):
    """Attach index list file to marking job"""
    user_id = 1  # TODO: Get from authentication
    try:
        # Get the marking job
        result = await db.execute(
            select(MarkingJob).where(
                MarkingJob.id == marking_job_id,
                MarkingJob.created_by == user_id
            )
        )
        marking = result.scalar_one_or_none()
        
        if not marking:
            raise HTTPException(status_code=404, detail="Marking job not found")
        
        # Update index list file ID
        marking.index_list_file_id = index_list_data.index_list_file_id
        await db.commit()
        await db.refresh(marking)
        
        logger.info(f"Index list attached to job {marking_job_id}")
        
        return _get_marking_response(marking)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to attach index list to job {marking_job_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to attach index list")

@router.post("/{marking_job_id}/start-marking", response_model=MarkingResponse)
@require_non_super_admin(require_admin_verified=True)
async def start_marking(
    request: Request,
    marking_job_id: int,
    db: AsyncSession = Depends(get_async_db)
):
    """Start the marking process"""
    # Get user from token for ownership validation
    user_id = request.state.current_user.id
    try:
        # Get the marking job
        result = await db.execute(
            select(MarkingJob).where(
                MarkingJob.id == marking_job_id,
                MarkingJob.created_by == user_id
            )
        )
        marking = result.scalar_one_or_none()
        
        if not marking:
            raise HTTPException(status_code=404, detail="Marking job not found")
        
        if marking.status not in [MarkingJobStatus.ANSWER_SHEETS_ATTACHED, MarkingJobStatus.COMPLETED, MarkingJobStatus.FAILED]:
            raise HTTPException(status_code=400, detail="Job must be in answer sheets attached status or higher to start")
        
        # Validate preconditions
        if not marking.marking_scheme_id:
            raise HTTPException(status_code=400, detail="Marking scheme must be attached before starting")
        
        if not marking.answer_sheets_folder_id:
            raise HTTPException(status_code=400, detail="Answer sheets must be attached before starting")
        
        # TODO: Check if template config job is completed (when applicable)
        # This would check the status of any associated TemplateConfigJob
        
        # Generate result sheet file path if not already set
        if not marking.result_sheet_file_path or marking.result_sheet_file_path == 'pending':
            random_id = str(uuid.uuid4())[:8]
            marking.result_sheet_file_path = f"results/{user_id}/{marking.id}_{random_id}_output.xlsx"
            await db.commit()
            await db.refresh(marking)
        
        try:
            logger.info(f"Starting marking job {marking_job_id}")
            await submit_marking_job(marking_job_id, db)
            marking.status = MarkingJobStatus.QUEUED
            await db.commit()
            await db.refresh(marking)
            logger.info(f"Successfully submitted marking job {marking_job_id} to queue")
        except Exception as e:
            logger.error(f"Failed to submit marking job {marking_job_id} to queue: {e}")
            # Revert status to pending on queue failure
            marking.status = MarkingJobStatus.FAILED
            marking.error_message = f"Failed to submit job to queue: {str(e)}"
            await db.commit()
            raise HTTPException(status_code=500, detail="Failed to submit job to processing queue")
        
        return _get_marking_response(marking)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to start marking job {marking_job_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to start marking job")




@router.get("/{marking_job_id}/results", response_model=ResultsData)
@require_non_super_admin(require_admin_verified=True)
async def get_results(
    request: Request,
    marking_job_id: int,
    db: AsyncSession = Depends(get_async_db)
):
    """Get the results for a marking job"""
    # Get user from token for ownership validation
    user_id = request.state.current_user.id
    try:
        # Get the marking job
        result = await db.execute(
            select(MarkingJob).options(selectinload(MarkingJob.template)).where(
                MarkingJob.id == marking_job_id,
                MarkingJob.created_by == user_id
            )
        )
        marking = result.scalar_one_or_none()
        
        if not marking:
            raise HTTPException(status_code=404, detail="Marking job not found")
        
        return ResultsData(
            id=marking.id,
            name=marking.name,
            status=marking.status,
            priority=marking.priority,
            template_name=marking.template.name,
            created_at=marking.created_at,
            updated_at=marking.updated_at,
            created_by=marking.created_by,
            save_intermediate_results=marking.save_intermediate_results,
            marking_config_id=marking.marking_config_id,
            result_sheet_file_id=marking.result_sheet_file_id,
            processing_started_at=marking.processing_started_at,
            processing_completed_at=marking.processing_completed_at
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get results for marking job {marking_job_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to get results")



@router.put("/{marking_job_id}/update-result/{row_number}", response_model=MarkingResponse)
@require_non_super_admin(require_admin_verified=True)
async def update_result(
    request_obj: Request,
    marking_job_id: int,
    row_number: int,
    request: UpdateResultRequest,
    db: AsyncSession = Depends(get_async_db)
):
    """Update a result for a marking job"""
    # Get user from token for ownership validation
    user_id = request_obj.state.current_user.id
    try:
        # Get the marking job
        result = await db.execute(
            select(MarkingJob).options(selectinload(MarkingJob.result_sheet_file)).where(
                MarkingJob.id == marking_job_id,
                MarkingJob.created_by == user_id
            )
        )
        marking = result.scalar_one_or_none()
        
        if not marking:
            raise HTTPException(status_code=404, detail="Marking job not found")
        
        if not marking.result_sheet_file:
            raise HTTPException(status_code=404, detail="Result sheet file not found")
        
        # Update the result in the Excel file
        result_sheet_file_path = marking.result_sheet_file.path
        storage = SharedStorage()
        
        # Get the current Excel file
        excel_content = await storage.get_file(result_sheet_file_path)
        
        # Load the workbook from bytes
        workbook = load_workbook(BytesIO(excel_content))
        worksheet = workbook.active
        student_result = request.result
        
        correct_str = ",".join(map(str, student_result.correct)) if student_result.correct else "-"
        incorrect_str = ",".join(map(str, student_result.incorrect)) if student_result.incorrect else "-"
        more_than_one_marked_str = ",".join(map(str, student_result.more_than_one_marked)) if student_result.more_than_one_marked else "-"
        not_marked_str = ",".join(map(str, student_result.not_marked)) if student_result.not_marked else "-"
        columnwise_total_str = ",".join(map(str, student_result.columnwise_total)) if student_result.columnwise_total else "-"
        
        # Convert labeled_points to JSON string
        labeled_points_json = json.dumps(student_result.labeled_points, default=lambda x: x.dict() if hasattr(x, 'dict') else x)
        
        # Update the row (row_number is 1-based in Excel, but we need to account for header row)
        excel_row = row_number + 1  # +1 because Excel is 1-based and we have a header row
        
        worksheet.cell(row=excel_row, column=1, value=student_result.index_number)  # A
        worksheet.cell(row=excel_row, column=2, value=correct_str)  # B
        worksheet.cell(row=excel_row, column=3, value=incorrect_str)  # C
        worksheet.cell(row=excel_row, column=4, value=more_than_one_marked_str)  # D
        worksheet.cell(row=excel_row, column=5, value=not_marked_str)  # E
        worksheet.cell(row=excel_row, column=6, value=columnwise_total_str)  # F
        worksheet.cell(row=excel_row, column=7, value=student_result.score)  # G
        worksheet.cell(row=excel_row, column=8, value=student_result.flag)  # H
        worksheet.cell(row=excel_row, column=9, value=student_result.flag_reason)  # I
        worksheet.cell(row=excel_row, column=10, value=student_result.answer_sheet_path)  # J
        worksheet.cell(row=excel_row, column=11, value=labeled_points_json)  # K
        
        # Save the workbook back to bytes
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        updated_excel_content = output.getvalue()
        
        # Save the updated file back to storage
        await storage.save_file(updated_excel_content, result_sheet_file_path)
        
        await db.commit()
        await db.refresh(marking)
        return _get_marking_response(marking)   
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update result for marking job {marking_job_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update result")
    
    

@router.get("/{marking_job_id}/download-results")
@require_non_super_admin(require_admin_verified=True)
async def get_file(
    request: Request,
    marking_job_id: int,
    db: AsyncSession = Depends(get_async_db)
):
    """
    Endpoint to retrieve a generated XLSX file,
    remove some columns in-memory, and return for download.
    """

    # Validate user ownership
    user_id = request.state.current_user.id

    try:
        # Get the marking job
        result = await db.execute(
            select(MarkingJob).options(selectinload(MarkingJob.result_sheet_file)).where(
                MarkingJob.id == marking_job_id,
                MarkingJob.created_by == user_id
            )
        )
        marking = result.scalar_one_or_none()

        if not marking:
            raise HTTPException(status_code=404, detail="Marking job not found")
        
        if not marking.result_sheet_file:
            raise HTTPException(status_code=404, detail="Result sheet file not found")
        
        # Update the result in the Excel file
        result_sheet_file_path = marking.result_sheet_file.path
        
        # Load file from your shared storage
        shared_storage = SharedStorage()
        file_content = await shared_storage.get_file(result_sheet_file_path)

        if not file_content:
            raise HTTPException(status_code=404, detail="File not found")

        # ---- Perform in-memory operations ----
        # Create an in-memory buffer for reading
        original_buffer = BytesIO(file_content)

        # Read XLSX into pandas DataFrame
        df = pd.read_excel(original_buffer)

        # Example: remove some columns safely if they exist
        columns_to_remove = ["Answer Sheet Path", "Labeled Points", "Audit File Name"]
        df = df.drop(columns=[c for c in columns_to_remove if c in df.columns])

        # Create another in-memory buffer to save modified file
        modified_buffer = BytesIO()
        df.to_excel(modified_buffer, index=False)
        modified_buffer.seek(0)

        # Set headers for download (proper filename)
        headers = {
            "Content-Disposition": f'attachment; filename="{marking.name}_results.xlsx"'
        }

        # Return modified file (no disk writes!)
        return Response(
            content=modified_buffer.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to retrieve or process file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")
    

@router.get("/{marking_job_id}/download-audit")
@require_non_super_admin(require_admin_verified=True)
async def get_audit(
    request: Request,
    marking_job_id: int,
    db: AsyncSession = Depends(get_async_db)
):
    """
    Endpoint to retrieve a generated Audit ZIP file.
    """

    # Validate user ownership
    user_id = request.state.current_user.id

    try:
        # Get the marking job
        result = await db.execute(
            select(MarkingJob).options(
                selectinload(MarkingJob.result_sheet_file),
                selectinload(MarkingJob.template).selectinload(Template.template_file),
                selectinload(MarkingJob.marking_scheme),
            ).where(
                MarkingJob.id == marking_job_id,
                MarkingJob.created_by == user_id
            )
        )
        marking = result.scalar_one_or_none()

        if not marking:
            raise HTTPException(status_code=404, detail="Marking job not found")
        
        if not marking.result_sheet_file:
            raise HTTPException(status_code=404, detail="Result sheet file not found")
        
        # Update the result in the Excel file
        result_sheet_file_path = marking.result_sheet_file.path
        
        # Load file from your shared storage
        shared_storage = SharedStorage()
        excel_file_content = await shared_storage.get_file(result_sheet_file_path)

        if not excel_file_content:
            raise HTTPException(status_code=404, detail="File not found")

        # ---- Perform in-memory operations ----
        # Create an in-memory buffer for reading
        original_buffer = BytesIO(excel_file_content)

        # Read XLSX into pandas DataFrame
        df = pd.read_excel(original_buffer)

        # Example: remove some columns safely if they exist
        columns_to_remove = ["Answer Sheet Path", "Labeled Points"]
        df = df.drop(columns=[c for c in columns_to_remove if c in df.columns])

        # Create another in-memory buffer to save modified file
        modified_buffer = BytesIO()
        df.to_excel(modified_buffer, index=False)
        modified_buffer.seek(0)

        # Create an in-memory ZIP containing the modified results + template + marking scheme
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            # Add modified results XLSX
            zf.writestr(f"{marking.name}_results.xlsx", modified_buffer.getvalue())

            # Add template files if available
            try:
                if getattr(marking, "template", None) and getattr(marking.template, "template_file", None) and getattr(marking.template.template_file, "path", None):
                    tpl_path = marking.template.template_file.path
                    tpl_bytes = await shared_storage.get_file(tpl_path)
                    if tpl_bytes:
                        zf.writestr(os.path.basename(tpl_path), tpl_bytes)
            except Exception:
                logger.warning(f"Failed to include template file for marking job {marking_job_id}")

            # Add marking scheme file if available
            try:
                if getattr(marking, "marking_scheme", None) and getattr(marking.marking_scheme, "path", None):
                    ms_path = marking.marking_scheme.path
                    ms_bytes = await shared_storage.get_file(ms_path)
                    if ms_bytes:
                        zf.writestr(os.path.basename(ms_path), ms_bytes)
            except Exception:
                logger.warning(f"Failed to include marking scheme file for marking job {marking_job_id}")
            
            #Add intermediate results folder
            try:
                inter_rel = getattr(marking, "intermediate_results_path", None)
                if inter_rel:
                    base_path = shared_storage.get_base_path()
                    inter_folder = base_path / inter_rel
                    if inter_folder.exists() and inter_folder.is_dir():
                        # Walk directory and add files preserving relative structure
                        for root, dirs, files in os.walk(inter_folder):
                            for fname in files:
                                file_full = Path(root) / fname
                                # relative path inside the intermediate folder
                                rel_inside = os.path.relpath(file_full, inter_folder)
                                arcname = os.path.join(os.path.basename(inter_rel.rstrip('/')), rel_inside)
                                try:
                                    # Use shared_storage to read file bytes
                                    file_path_for_storage = str(Path(inter_rel) / rel_inside)
                                    file_bytes = await shared_storage.get_file(file_path_for_storage)
                                except Exception:
                                    # fallback to reading directly from disk
                                    try:
                                        with open(file_full, 'rb') as f:
                                            file_bytes = f.read()
                                    except Exception:
                                        file_bytes = None

                                if file_bytes:
                                    zf.writestr(arcname, file_bytes)
            except Exception:
                logger.warning(f"Failed to include intermediate results folder for marking job {marking_job_id}")

        zip_buffer.seek(0)

        headers = {"Content-Disposition": f'attachment; filename="{marking.name}_audit.zip"'}

        # Return ZIP (no disk writes)
        return Response(
            content=zip_buffer.read(),
            media_type="application/zip",
            headers=headers,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to retrieve or process file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")